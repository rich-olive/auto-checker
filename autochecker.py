# open an excel spreadsheet

# later this can be done using user input and a path

# cycle through each column

# check content against criteria, using logic

# logic including regex checking for forbidden characters, certain characters, whitespace
# check for length

import openpyxl
from pathlib import Path
import re
from openpyxl.styles import Font, PatternFill

path = Path.home()/'Desktop'/'template_dummy.xlsx'
# "C:\Users\olivi\Desktop\template_dummy.xlsx"
print(path)

wb = openpyxl.load_workbook(path)

ws = wb["Catalogue Template"]
print(ws.title)

# department_code = ws["A4"]
# print(department_code.value)
# department_code_range = "A4:A"+str()

# max_col = inclusive
test_string = "hello!!2"
alphanumeric_regex = r"\[A-Z]"
error_fill =PatternFill(start_color="0000FF", end_color="0000FF", patternType="solid")

for col in ws.iter_cols(min_row=4, max_col=1):
    for cell in col:
        print(cell.value)
        # print(type(cell.value))
        if type(cell.value) != str:
            cell.fill = error_fill
        # if not bool(re.match(alphanumeric_regex, cell.value)):
        #     print("not a capital letter!")

# result = bool(re.match(alphanumeric_regex,test_string))
# print(result)

wb.save(path)